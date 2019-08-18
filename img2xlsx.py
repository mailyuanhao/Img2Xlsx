from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws = wb.active
im = Image.open("1.png")
s = im.size
im = im.resize((int(s[0] / 2), int(s[1] / 1)), Image.ANTIALIAS)
s = im.size
px = im.load()
for x in range(0, s[0]):
    for y in range(0, s[1]):
        pi = px[x, y]
        c = "{:0>2X}{:0>2X}{:0>2X}".format(pi[0], pi[1], pi[2])
        d = ws.cell(column=x + 1, row=y + 1)
        d.fill = PatternFill('solid', fgColor=c)

wb.save("abc.xlsx")